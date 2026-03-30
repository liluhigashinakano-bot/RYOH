"""12月ファイルの診断"""
import sys, io
sys.path.insert(0, '.')
import openpyxl
from app.routers.excel_import import parse_daily_sheets, _calc_prefs_from_monthly

path = r"C:\Users\lalal\Downloads\東中野PC日報12月2025.xlsx"
day_labels = ["月", "火", "水", "木", "金", "土", "日"]

with open(path, "rb") as f:
    content = f.read()

wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
print("シート一覧:", wb.sheetnames[:5])

# 年月自動検出
year, month = 2025, 12
try:
    if "1日" in wb.sheetnames:
        first_row = list(wb["1日"].iter_rows(values_only=True))[0]
        print(f"1行目[17]={first_row[17]}, [21]={first_row[21]}")
        if first_row[17] and isinstance(first_row[17], (int, float)):
            year = int(first_row[17])
        if first_row[21] and isinstance(first_row[21], (int, float)):
            month = int(first_row[21])
except Exception as e:
    print(f"年月検出エラー: {e}")

print(f"検出年月: {year}年{month}月")

try:
    visits = parse_daily_sheets(wb, year, month)
    print(f"総来店レコード: {len(visits)}")
except Exception as e:
    print(f"parse_daily_sheets エラー: {e}")
    import traceback; traceback.print_exc()
    sys.exit(1)

# 顧客別集計テスト
from collections import defaultdict
seen = set()
errors = []
for v in visits:
    name = v["customer_name"]
    if name in seen: continue
    seen.add(name)
    cvs = [x for x in visits if x["customer_name"] == name]

    month_key = f"{year}-{month:02d}"
    this_month_data = {
        "visits": len(cvs),
        "spend": sum(x["total_payment"] for x in cvs),
        "extensions": sum(x["extensions"] for x in cvs),
        "persons": sum(x["group_size"] for x in cvs),
        "set_l": sum(x["set_l"] for x in cvs),
        "set_mg": sum(x["set_mg"] for x in cvs),
        "set_shot": sum(x.get("set_shot", 0) for x in cvs),
        "in_mins": [(t // 100) * 60 + (t % 100) for x in cvs if (t := x["in_time"])],
        "arrival_source": {},
        "day_prefs": {},
    }
    for x in cvs:
        src = x["arrival_source"] if x["arrival_source"] else "不明"
        this_month_data["arrival_source"][src] = this_month_data["arrival_source"].get(src, 0) + 1

    src_total = sum(this_month_data["arrival_source"].values())
    if src_total != len(cvs):
        errors.append(f"[{name}] 動機合計{src_total} != 来店{len(cvs)}")

    try:
        monthly_data = {month_key: this_month_data}
        _calc_prefs_from_monthly(monthly_data, day_labels)
    except Exception as e:
        errors.append(f"[{name}] calc error: {e}")

if errors:
    print(f"\nエラー {len(errors)}件:")
    for e in errors[:10]: print(f"  {e}")
else:
    print(f"全テストパス ({len(seen)}名)")
