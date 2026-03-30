"""旧型式Excelインポートの単体テスト"""
import sys, json
sys.path.insert(0, '.')
import openpyxl, io
from app.routers.excel_import import parse_daily_sheets

path = r"C:\Users\lalal\Downloads\東中野PC日報2月2026.xlsx"
with open(path, "rb") as f:
    content = f.read()

wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

# 年月自動検出
year, month = 2026, 2
try:
    if "1日" in wb.sheetnames:
        first_row = list(wb["1日"].iter_rows(values_only=True))[0]
        if first_row[17] and isinstance(first_row[17], (int, float)):
            year = int(first_row[17])
        if first_row[21] and isinstance(first_row[21], (int, float)):
            month = int(first_row[21])
except Exception as e:
    print(f"年月検出エラー: {e}")

print(f"年月: {year}年{month}月")

visits = parse_daily_sheets(wb, year, month)
print(f"総来店レコード数: {len(visits)}")

# 田口の来店データを表示
taguchi = [v for v in visits if v["customer_name"] == "田口"]
print(f"\n田口の来店数: {len(taguchi)}")
for v in taguchi:
    print(f"  {v}")

# 集計テスト（田口）
if taguchi:
    from datetime import date
    visit_count = len(taguchi)
    total_spend = sum(v["total_payment"] for v in taguchi)
    avg_spend = int(total_spend / visit_count)
    total_extensions = sum(v["extensions"] for v in taguchi)
    avg_extensions = round(total_extensions / visit_count, 2)
    in_times = [v["in_time"] for v in taguchi if v["in_time"]]
    avg_in_time = int(sum(in_times) / len(in_times)) if in_times else None
    print(f"\n田口 集計:")
    print(f"  total_spend={total_spend}, avg_spend={avg_spend}")
    print(f"  total_extensions={total_extensions}, avg_extensions={avg_extensions}")
    print(f"  avg_in_time={avg_in_time}")
