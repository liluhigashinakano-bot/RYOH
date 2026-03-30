"""
Excelファイル（東中野PC日報形式）から顧客データをインポートするAPI
"""
import os
import re
import io
import csv
from datetime import datetime, date
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified
from pydantic import BaseModel

from ..database import get_db
from .. import models
from ..auth import get_current_user

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

router = APIRouter(prefix="/api/excel", tags=["excel-import"])

IMPORTS_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "data", "imports")


def ensure_imports_dir():
    os.makedirs(IMPORTS_DIR, exist_ok=True)


# ─── 顧客情報一覧シート パーサー ──────────────────────────────────────────────

def parse_customer_list_sheet(ws) -> list[dict]:
    """顧客情報一覧シートをパース"""
    customers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4:
            continue
        if row[0] is None or not isinstance(row[0], (int, float)):
            continue
        name = row[1]
        if not name or str(name).strip() == "":
            continue

        last_visit = row[2]
        last_visit_date = None
        if isinstance(last_visit, datetime):
            last_visit_date = last_visit.date()
        elif isinstance(last_visit, date):
            last_visit_date = last_visit

        birthday_val = row[28] if len(row) > 28 else None
        birthday = None
        if isinstance(birthday_val, datetime):
            birthday = birthday_val.date()
        elif isinstance(birthday_val, date):
            birthday = birthday_val

        assigned_casts = []
        for ci in range(29, min(33, len(row))):
            if row[ci] and str(row[ci]).strip():
                assigned_casts.append(str(row[ci]).strip())

        arrival_map = {
            "ティッシュ": row[19] if len(row) > 19 else 0,
            "アメブロ": row[20] if len(row) > 20 else 0,
            "LINE": row[21] if len(row) > 21 else 0,
            "紹介": row[22] if len(row) > 22 else 0,
            "SNS": row[23] if len(row) > 23 else 0,
            "Google": row[24] if len(row) > 24 else 0,
            "看板": row[25] if len(row) > 25 else 0,
            "その他": row[26] if len(row) > 26 else 0,
        }
        arrival_source = {k: int(v) for k, v in arrival_map.items() if v and isinstance(v, (int, float)) and v > 0}

        day_labels = ["月", "火", "水", "木", "金", "土", "日"]
        day_prefs = {}
        for di, label in enumerate(day_labels):
            val = row[12 + di] if len(row) > 12 + di else 0
            if val and isinstance(val, (int, float)) and val > 0:
                day_prefs[label] = int(val)

        def safe_num(v, default=0):
            if v is None or not isinstance(v, (int, float)):
                return default
            return v

        customers.append({
            "name": str(name).strip(),
            "last_visit_date": last_visit_date,
            "total_visits": int(safe_num(row[6])),
            "total_spend": int(safe_num(row[4])),
            "birthday": birthday,
            "preferences": {
                "avg_spend": int(safe_num(row[5])),
                "avg_extensions": round(float(safe_num(row[7])), 2),
                "avg_duration_min": int(safe_num(row[8])),
                "set_l": round(float(safe_num(row[9])), 2),
                "set_mg": round(float(safe_num(row[10])), 2),
                "set_shot": round(float(safe_num(row[11])), 2),
                "day_prefs": day_prefs,
                "arrival_source": arrival_source,
                "avg_group_size": round(float(safe_num(row[27])), 1),
                "assigned_casts": assigned_casts,
            },
        })
    return customers


# ─── 日別シート パーサー ─────────────────────────────────────────────────────

def _to_num(v, default=0):
    """ExcelセルのInt/Float/文字列を数値に変換"""
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        cleaned = v.replace(',', '').replace('，', '').strip()
        try:
            return float(cleaned)
        except ValueError:
            return default
    return default


def _serialize_cell(v):
    """ExcelセルをJSON保存可能な値に変換"""
    if v is None:
        return None
    if isinstance(v, (int, float, str, bool)):
        return v
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return str(v)


def extract_section_visits(rows: list, start_row: int, end_row: int,
                            is_repeat: bool, visit_date_str: str,
                            col_headers: list | None = None) -> list[dict]:
    """リピーター/新規セクションから来店データを抽出"""
    section_visits = []
    i = start_row
    while i < end_row and i < len(rows):
        row = rows[i]
        if row[0] is not None and isinstance(row[0], (int, float)) and row[0] > 0:
            in_val = row[1] if len(row) > 1 else None
            out_val = row[4] if len(row) > 4 else None
            cash_val = _to_num(row[27] if len(row) > 27 else None)
            card_val = _to_num(row[32] if len(row) > 32 else None)
            if not in_val and not out_val and cash_val == 0 and card_val == 0:
                i += 1
                continue
            customer_raw = row[40] if len(row) > 40 else None
            if customer_raw:
                names = [n.strip() for n in re.split(r'[、,，・]', str(customer_raw)) if n.strip()]
                cash = _to_num(row[27] if len(row) > 27 else 0)
                card = _to_num(row[32] if len(row) > 32 else 0)
                total = cash + card
                in_t = row[1] if len(row) > 1 else None
                out_t = row[4] if len(row) > 4 else None
                # B列(index1)〜AU列(index46)の生データをdict形式で保存
                raw_cells = list(row[1:47]) if len(row) >= 47 else list(row[1:]) + [None] * (46 - len(row) + 1)
                if col_headers:
                    raw_data = {col_headers[i]: _serialize_cell(raw_cells[i]) for i in range(len(col_headers))}
                else:
                    raw_data = {f"col{i+2}": _serialize_cell(v) for i, v in enumerate(raw_cells)}

                for name in names:
                    section_visits.append({
                        "date": visit_date_str,
                        "customer_name": name,
                        "is_repeat": is_repeat,
                        "table_no": str(row[23]) if len(row) > 23 and row[23] else str(int(row[0])),
                        "in_time": int(_to_num(in_t)) if in_t else None,
                        "out_time": int(_to_num(out_t)) if out_t else None,
                        "extensions": int(_to_num(row[7] if len(row) > 7 else 0)),
                        "course": str(row[12]).lower() if len(row) > 12 and row[12] else "",
                        "set_l": float(_to_num(row[16] if len(row) > 16 else 0)),
                        "set_shot": float(_to_num(row[17] if len(row) > 17 else 0)),
                        "set_mg": float(_to_num(row[18] if len(row) > 18 else 0)),
                        "group_size": max(1, int(_to_num(row[25] if len(row) > 25 else 1, default=1))),
                        "payment_cash": int(cash),
                        "payment_card": int(card),
                        "total_payment": int(total),
                        "arrival_source": str(row[43]) if len(row) > 43 and row[43] else "",
                        "raw_data": raw_data,
                    })
        i += 1
    return section_visits


def parse_daily_sheets(wb, year: int, month: int) -> list[dict]:
    """日別シート（1日〜31日）から来店データを抽出"""
    visits = []
    for day in range(1, 32):
        sheet_name = f"{day}日"
        if sheet_name not in wb.sheetnames:
            continue
        try:
            visit_date = date(year, month, day)
        except ValueError:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        visit_date_str = visit_date.isoformat()

        # B〜AU列のヘッダーをrow[5]から取得（index5=6行目がヘッダー行）
        col_headers = None
        if len(rows) > 5:
            hrow = rows[5]
            headers = [str(hrow[i]).strip() if hrow[i] is not None else f"col{i+2}" for i in range(1, 47)]
            col_headers = headers

        new_section_start = None
        for i, row in enumerate(rows):
            if row[0] == "新規":
                new_section_start = i
                break

        repeat_end = new_section_start if new_section_start is not None else len(rows)
        visits += extract_section_visits(rows, 6, repeat_end, True, visit_date_str, col_headers)
        if new_section_start is not None:
            visits += extract_section_visits(rows, new_section_start + 2, len(rows), False, visit_date_str, col_headers)

    return visits


# ─── 月別データから集計値を再計算 ────────────────────────────────────────────

def _calc_prefs_from_monthly(monthly_data: dict, day_labels: list) -> dict:
    """monthly_dataの全月分から集計値を再計算する"""
    total_visits = sum(m["visits"] for m in monthly_data.values())
    total_spend = sum(m["spend"] for m in monthly_data.values())
    total_extensions = sum(m["extensions"] for m in monthly_data.values())
    total_persons = sum(m["persons"] for m in monthly_data.values())
    total_set_l = sum(m["set_l"] for m in monthly_data.values())
    total_set_mg = sum(m["set_mg"] for m in monthly_data.values())
    total_set_shot = sum(m["set_shot"] for m in monthly_data.values())
    all_in_mins = [mn for m in monthly_data.values() for mn in m.get("in_mins", [])]

    avg_spend = int(total_spend / total_visits) if total_visits > 0 else 0
    avg_extensions = round(total_extensions / max(total_persons, 1), 2)
    avg_group = round(total_persons / total_visits, 1) if total_visits > 0 else 1
    divisor = total_extensions + 1
    set_l_avg = round(total_set_l / divisor, 2)
    set_mg_avg = round(total_set_mg / divisor, 2)
    set_shot_avg = round(total_set_shot / divisor, 2)

    if all_in_mins:
        avg_min = int(sum(all_in_mins) / len(all_in_mins))
        avg_in_time = (avg_min // 60) * 100 + (avg_min % 60)
    else:
        avg_in_time = None

    # 月間平均来店 = 合計来店数 ÷ 月数（重複なし）
    monthly_avg_visits = round(total_visits / max(len(monthly_data), 1), 1)

    # 来店動機を全月でマージ
    merged_src: dict = {}
    for m in monthly_data.values():
        for k, cnt in m.get("arrival_source", {}).items():
            merged_src[k] = merged_src.get(k, 0) + cnt

    # 曜日別来店数を全月でマージ
    merged_day: dict = {}
    for m in monthly_data.values():
        for k, cnt in m.get("day_prefs", {}).items():
            merged_day[k] = merged_day.get(k, 0) + cnt

    # テスト: 来店動機合計 == 合計来店数
    src_total = sum(merged_src.values())
    assert src_total == total_visits, f"来店動機合計({src_total}) != 合計来店数({total_visits})"

    return {
        "_total_visits": total_visits,
        "_total_spend": total_spend,
        "avg_spend": avg_spend,
        "avg_extensions": avg_extensions,
        "avg_group_size": avg_group,
        "avg_in_time": avg_in_time,
        "monthly_avg_visits": monthly_avg_visits,
        "set_l": set_l_avg,
        "set_mg": set_mg_avg,
        "set_shot": set_shot_avg,
        "day_prefs": merged_day,
        "arrival_source": merged_src,
    }


# ─── Pydanticモデル ──────────────────────────────────────────────────────────

class ImportResult(BaseModel):
    created: int
    updated: int
    skipped: int
    file_saved: str
    customers: list[str]


class DailyImportResult(BaseModel):
    visits_extracted: int
    customers_created: int
    customers_updated: int
    csv_saved: str
    store_name: str


# ─── エンドポイント ──────────────────────────────────────────────────────────

@router.post("/import-customers", response_model=ImportResult)
async def import_customers_from_excel(
    file: UploadFile = File(...),
    store_name: str = Form(default="東中野"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """顧客情報一覧シートから顧客を一括登録"""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxlがインストールされていません")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Excelファイル(.xlsx)をアップロードしてください")

    ensure_imports_dir()
    content = await file.read()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    save_name = f"{timestamp}_{store_name}_{file.filename}"
    save_path = os.path.join(IMPORTS_DIR, save_name)
    with open(save_path, "wb") as f:
        f.write(content)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excelファイルの読み込みに失敗: {str(e)}")

    target_sheet = None
    for sname in wb.sheetnames:
        if "顧客情報一覧" in sname:
            target_sheet = wb[sname]
            break
    if target_sheet is None:
        raise HTTPException(status_code=400, detail="「顧客情報一覧」シートが見つかりません")

    customers_data = parse_customer_list_sheet(target_sheet)
    created = updated = skipped = 0
    imported_names = []

    for cdata in customers_data:
        name = cdata["name"]
        if not name:
            skipped += 1
            continue
        existing = db.query(models.Customer).filter(
            models.Customer.name == name,
            models.Customer.is_active == True,
        ).first()
        if existing:
            existing.last_visit_date = cdata["last_visit_date"] or existing.last_visit_date
            existing.total_visits = max(existing.total_visits, cdata["total_visits"])
            existing.total_spend = max(existing.total_spend, cdata["total_spend"])
            if cdata["birthday"] and not existing.birthday:
                existing.birthday = cdata["birthday"]
            existing.preferences = {**(existing.preferences or {}), **cdata["preferences"]}
            flag_modified(existing, "preferences")
            updated += 1
        else:
            customer = models.Customer(
                name=name,
                last_visit_date=cdata["last_visit_date"],
                total_visits=cdata["total_visits"],
                total_spend=cdata["total_spend"],
                birthday=cdata["birthday"],
                preferences=cdata["preferences"],
            )
            db.add(customer)
            created += 1
        imported_names.append(name)

    db.commit()
    return ImportResult(created=created, updated=updated, skipped=skipped,
                        file_saved=save_name, customers=imported_names)


@router.post("/import-daily-sheets", response_model=DailyImportResult)
async def import_daily_sheets(
    file: UploadFile = File(...),
    store_name: str = Form(default="東中野"),
    year: int = Form(default=2026),
    month: int = Form(default=2),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """日別シート（1日〜31日）から来店データを抽出してCSV保存・顧客登録"""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxlがインストールされていません")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Excelファイル(.xlsx)をアップロードしてください")

    ensure_imports_dir()
    content = await file.read()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excelの読み込みに失敗: {str(e)}")

    # 年月をExcelから自動検出
    try:
        if "1日" in wb.sheetnames:
            first_row = list(wb["1日"].iter_rows(values_only=True))[0]
            if first_row[17] and isinstance(first_row[17], (int, float)):
                year = int(first_row[17])
            if first_row[21] and isinstance(first_row[21], (int, float)):
                month = int(first_row[21])
    except Exception:
        pass

    visits = parse_daily_sheets(wb, year, month)

    # 店舗別フォルダにCSV保存
    store_dir = os.path.join(IMPORTS_DIR, store_name)
    os.makedirs(store_dir, exist_ok=True)
    csv_filename = f"{year}{month:02d}_visits.csv"
    csv_path = os.path.join(store_dir, csv_filename)

    fieldnames = ["date", "customer_name", "is_repeat", "table_no",
                  "in_time", "out_time", "extensions", "course",
                  "set_l", "set_shot", "set_mg", "group_size",
                  "payment_cash", "payment_card", "total_payment", "arrival_source"]
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(visits)

    # 顧客DBに登録・更新
    created = updated = 0
    seen_names: set[str] = set()

    day_labels = ["月", "火", "水", "木", "金", "土", "日"]

    for v in visits:
        name = v["customer_name"]
        if not name or name in seen_names:
            continue
        seen_names.add(name)

        customer_visits = [x for x in visits if x["customer_name"] == name]
        visit_dates = sorted([date.fromisoformat(x["date"]) for x in customer_visits])

        # 今回のファイルの月キー（YYYY-MM）
        month_key = f"{year}-{month:02d}"

        # 今回ファイル分の月別集計データ
        this_month_data = {
            "visits": len(customer_visits),
            "spend": sum(x["total_payment"] for x in customer_visits),
            "extensions": sum(x["extensions"] for x in customer_visits),
            "persons": sum(x["group_size"] for x in customer_visits),
            "set_l": sum(x["set_l"] for x in customer_visits),
            "set_mg": sum(x["set_mg"] for x in customer_visits),
            "set_shot": sum(x.get("set_shot", 0) for x in customer_visits),
            "in_mins": [(t // 100) * 60 + (t % 100) for x in customer_visits if (t := x["in_time"])],
            "arrival_source": {},
            "day_prefs": {},
        }
        # 来店動機（空は"不明"）
        for x in customer_visits:
            src = x["arrival_source"] if x["arrival_source"] else "不明"
            this_month_data["arrival_source"][src] = this_month_data["arrival_source"].get(src, 0) + 1
        # 曜日別
        for vd in visit_dates:
            label = day_labels[vd.weekday()]
            this_month_data["day_prefs"][label] = this_month_data["day_prefs"].get(label, 0) + 1

        existing = db.query(models.Customer).filter(
            models.Customer.name == name,
            models.Customer.is_active == True,
        ).first()

        if existing:
            old_prefs = existing.preferences or {}
            # monthly_data: 月別データを保持（同じ月を再インポートしたら上書き）
            monthly_data: dict = old_prefs.get("monthly_data", {})
            monthly_data[month_key] = this_month_data

            # 全月分から再集計
            all_prefs = _calc_prefs_from_monthly(monthly_data, day_labels)
            # 手動設定項目は保持
            for keep_key in ["ng_notes", "anniversary_date", "assigned_casts"]:
                if keep_key in old_prefs:
                    all_prefs[keep_key] = old_prefs[keep_key]
            all_prefs["monthly_data"] = monthly_data
            existing.preferences = all_prefs
            flag_modified(existing, "preferences")

            # first/last_visit_date も全月から算出
            fv = visit_dates[0] if visit_dates else None
            lv = visit_dates[-1] if visit_dates else None
            if fv and (not existing.first_visit_date or fv < existing.first_visit_date):
                existing.first_visit_date = fv
            if lv and (not existing.last_visit_date or lv > existing.last_visit_date):
                existing.last_visit_date = lv
            existing.total_visits = all_prefs["_total_visits"]
            existing.total_spend = all_prefs["_total_spend"]
            customer_obj = existing
            updated += 1
        else:
            monthly_data = {month_key: this_month_data}
            all_prefs = _calc_prefs_from_monthly(monthly_data, day_labels)
            all_prefs["monthly_data"] = monthly_data
            customer_obj = models.Customer(
                name=name,
                total_visits=all_prefs["_total_visits"],
                total_spend=all_prefs["_total_spend"],
                last_visit_date=visit_dates[-1] if visit_dates else None,
                first_visit_date=visit_dates[0] if visit_dates else None,
                preferences=all_prefs,
            )
            db.add(customer_obj)
            db.flush()  # IDを確定
            created += 1

        # 来店履歴をCustomerVisitに保存（同月同店舗は削除→再挿入）
        if customer_obj.id and visit_dates:
            db.query(models.CustomerVisit).filter(
                models.CustomerVisit.customer_id == customer_obj.id,
                models.CustomerVisit.store_name == store_name,
                models.CustomerVisit.date >= visit_dates[0],
                models.CustomerVisit.date <= visit_dates[-1],
            ).delete(synchronize_session=False)
        if customer_obj.id:
            for v in customer_visits:
                db.add(models.CustomerVisit(
                    customer_id=customer_obj.id,
                    date=date.fromisoformat(v["date"]),
                    store_name=store_name,
                    is_repeat=v["is_repeat"],
                    in_time=v.get("in_time"),
                    out_time=v.get("out_time"),
                    total_payment=v.get("total_payment", 0),
                    raw_data=v.get("raw_data", {}),
                ))

    db.commit()
    return DailyImportResult(
        visits_extracted=len(visits),
        customers_created=created,
        customers_updated=updated,
        csv_saved=f"{store_name}/{csv_filename}",
        store_name=store_name,
    )


@router.get("/imports")
def list_imports(current_user: models.User = Depends(get_current_user)):
    """保存済みインポートファイル一覧"""
    ensure_imports_dir()
    files = []
    for fname in sorted(os.listdir(IMPORTS_DIR), reverse=True):
        fpath = os.path.join(IMPORTS_DIR, fname)
        stat = os.stat(fpath)
        files.append({
            "filename": fname,
            "size_kb": round(stat.st_size / 1024, 1),
            "imported_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
        })
    return files
